from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg, Count, Min, Max
from django.db.models.functions import ExtractYear
from django.http import HttpResponse

from .models import SupplierEvaluation, BuyerEvaluation
from .forms import SupplierEvaluationForm, BuyerEvaluationForm
from suppliers.models import Supplier


@login_required
def evaluation_list(request):
    """Liste des évaluations"""
    evaluations = SupplierEvaluation.objects.select_related('supplier', 'evaluator').all()
    
    # Filtrage par fournisseur
    supplier_id = request.GET.get('supplier')
    include_yearly = request.GET.get('yearly') in ['1', 'true', 'on']
    include_chart = request.GET.get('chart') in ['1', 'true', 'on']
    if supplier_id:
        evaluations = evaluations.filter(supplier_id=supplier_id)
    
    # Recherche
    search = request.GET.get('search')
    if search:
        evaluations = evaluations.filter(
            Q(supplier__nom_complet_organisation__icontains=search) |
            Q(evaluator__email__icontains=search) |
            Q(comments__icontains=search)
        )
    
    # Statistiques
    stats = {
        'total': evaluations.count(),
        'avg_rating': evaluations.aggregate(Avg('vendor_final_rating'))['vendor_final_rating__avg'] or 0,
    }
    
    # Liste des fournisseurs pour le filtre
    suppliers = Supplier.objects.filter(actif=True).order_by('nom_complet_organisation')
    
    context = {
        'evaluations': evaluations,
        'suppliers': suppliers,
        'stats': stats,
    }
    return render(request, 'evaluations/evaluation_list.html', context)


@login_required
def evaluation_detail(request, pk):
    """Détail d'une évaluation"""
    evaluation = get_object_or_404(SupplierEvaluation, pk=pk)
    
    # Récupérer toutes les évaluations du même fournisseur
    supplier_evaluations = SupplierEvaluation.objects.filter(
        supplier=evaluation.supplier
    ).exclude(pk=pk).order_by('-date_evaluation')[:5]
    
    # Préparer les descriptions des critères pour le template
    descriptions = {
        'delivery_compliance': evaluation.get_criteria_description('delivery_compliance', evaluation.delivery_compliance),
        'delivery_timeline': evaluation.get_criteria_description('delivery_timeline', evaluation.delivery_timeline),
        'advising_capability': evaluation.get_criteria_description('advising_capability', evaluation.advising_capability),
        'after_sales_qos': evaluation.get_criteria_description('after_sales_qos', evaluation.after_sales_qos),
        'vendor_relationship': evaluation.get_criteria_description('vendor_relationship', evaluation.vendor_relationship),
    }
    
    context = {
        'evaluation': evaluation,
        'supplier_evaluations': supplier_evaluations,
        'descriptions': descriptions,
    }
    return render(request, 'evaluations/evaluation_detail.html', context)


@login_required
def evaluation_create(request):
    """Créer une évaluation"""
    if request.method == 'POST':
        form = SupplierEvaluationForm(request.POST)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.evaluator = request.user
            evaluation.save()
            messages.success(request, f'Évaluation du fournisseur "{evaluation.supplier.nom_complet_organisation}" créée avec succès !')
            return redirect('evaluations:detail', pk=evaluation.pk)
    else:
        # Pré-remplir le fournisseur si passé en paramètre
        supplier_id = request.GET.get('supplier')
        initial = {}
        if supplier_id:
            initial['supplier'] = supplier_id
        form = SupplierEvaluationForm(initial=initial)
    
    context = {'form': form}
    return render(request, 'evaluations/evaluation_form.html', context)


@login_required
def evaluation_edit(request, pk):
    """Modifier une évaluation"""
    evaluation = get_object_or_404(SupplierEvaluation, pk=pk)
    
    if request.method == 'POST':
        form = SupplierEvaluationForm(request.POST, instance=evaluation)
        if form.is_valid():
            evaluation = form.save()
            messages.success(request, f'Évaluation modifiée avec succès !')
            return redirect('evaluations:detail', pk=evaluation.pk)
    else:
        form = SupplierEvaluationForm(instance=evaluation)
    
    context = {'form': form, 'evaluation': evaluation}
    return render(request, 'evaluations/evaluation_form.html', context)


@login_required
def evaluation_delete(request, pk):
    """Supprimer une évaluation"""
    evaluation = get_object_or_404(SupplierEvaluation, pk=pk)
    
    if request.method == 'POST':
        supplier_name = evaluation.supplier.nom_complet_organisation
        evaluation.delete()
        messages.success(request, f'Évaluation du fournisseur "{supplier_name}" supprimée avec succès.')
        return redirect('evaluations:list')
    
    # Si GET, rediriger vers la liste (la suppression se fera via AJAX)
    return redirect('evaluations:list')


@login_required
def supplier_evaluations(request, supplier_id):
    """Liste des évaluations d'un fournisseur spécifique"""
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    evaluations = SupplierEvaluation.objects.filter(supplier=supplier).order_by('-date_evaluation')
    
    # Statistiques du fournisseur
    stats = {
        'total': evaluations.count(),
        'avg_rating': evaluations.aggregate(Avg('vendor_final_rating'))['vendor_final_rating__avg'] or 0,
        'avg_delivery_compliance': evaluations.aggregate(Avg('delivery_compliance'))['delivery_compliance__avg'] or 0,
        'avg_delivery_timeline': evaluations.aggregate(Avg('delivery_timeline'))['delivery_timeline__avg'] or 0,
        'avg_advising_capability': evaluations.aggregate(Avg('advising_capability'))['advising_capability__avg'] or 0,
        'avg_after_sales_qos': evaluations.aggregate(Avg('after_sales_qos'))['after_sales_qos__avg'] or 0,
        'avg_vendor_relationship': evaluations.aggregate(Avg('vendor_relationship'))['vendor_relationship__avg'] or 0,
    }
    
    context = {
        'supplier': supplier,
        'evaluations': evaluations,
        'stats': stats,
    }
    return render(request, 'evaluations/supplier_evaluations.html', context)


@login_required
def ranking_overview(request):
    """Supplier Ranking overview + supplier drilldown (legacy-like workflow)"""
    # Annoter avec les moyennes vendor et buyer
    suppliers_qs = Supplier.objects.annotate(
        avg_vendor_rating=Avg('evaluations__vendor_final_rating'),
        avg_buyer_rating=Avg('buyer_evaluations__buyer_final_rating'),
        vendor_eval_count=Count('evaluations', distinct=True),
        buyer_eval_count=Count('buyer_evaluations', distinct=True)
    ).filter(Q(vendor_eval_count__gt=0) | Q(buyer_eval_count__gt=0))

    total_suppliers = suppliers_qs.count()

    # Calculer les notes pondérées pour chaque fournisseur
    suppliers_with_weighted = []
    for supplier in suppliers_qs:
        weighted_rating = supplier.get_weighted_rating()
        suppliers_with_weighted.append({
            'id': supplier.id,
            'nom_complet_organisation': supplier.nom_complet_organisation,
            'avg_vendor_rating': supplier.avg_vendor_rating or 0,
            'avg_buyer_rating': supplier.avg_buyer_rating or 0,
            'weighted_rating': float(weighted_rating),
            'vendor_eval_count': supplier.vendor_eval_count,
            'buyer_eval_count': supplier.buyer_eval_count,
            'total_eval_count': supplier.vendor_eval_count + supplier.buyer_eval_count,
        })
    
    # Trier par note pondérée
    suppliers_with_weighted.sort(key=lambda x: x['weighted_rating'], reverse=True)
    
    # Ajouter les rangs
    for idx, s in enumerate(suppliers_with_weighted, start=1):
        s['rank'] = idx
    
    # Top 10 et Bottom 10
    top10 = suppliers_with_weighted[:10]
    bottom10 = list(reversed(suppliers_with_weighted[-10:]))
    
    # Liste pour le sélecteur (triée par nom)
    suppliers_stats = sorted(suppliers_with_weighted, key=lambda x: x['nom_complet_organisation'])

    # Selected supplier drilldown
    selected_supplier_id = request.GET.get('supplier')
    selected_supplier = None
    selected_supplier_data = None
    yearly_stats_list = []
    yearly_stats_json = []
    yearly_buyer_stats_list = []
    yearly_buyer_stats_json = []
    yearly_weighted_json = []
    chart_eval_labels = []
    chart_eval_values = []
    chart_eval_mavg = []
    evals_for_table = []

    if selected_supplier_id:
        selected_supplier = get_object_or_404(Supplier, pk=selected_supplier_id)
        
        # Évaluations vendor
        vendor_evals = SupplierEvaluation.objects.filter(supplier=selected_supplier)
        vendor_agg = vendor_evals.aggregate(
            avg_final=Avg('vendor_final_rating'),
            num=Count('id'),
            avg_delivery_compliance=Avg('delivery_compliance'),
            avg_delivery_timeline=Avg('delivery_timeline'),
            avg_advising_capability=Avg('advising_capability'),
            avg_after_sales_qos=Avg('after_sales_qos'),
            avg_vendor_relationship=Avg('vendor_relationship'),
        )
        
        # Évaluations acheteur
        buyer_evals = BuyerEvaluation.objects.filter(supplier=selected_supplier)
        buyer_agg = buyer_evals.aggregate(
            avg_final=Avg('buyer_final_rating'),
            num=Count('id'),
            avg_price_flexibility=Avg('price_flexibility'),
            avg_rfx_deadline_compliance=Avg('rfx_deadline_compliance'),
            avg_advisory_capability=Avg('advisory_capability'),
            avg_relationship_quality=Avg('relationship_quality'),
            avg_rfx_response_quality=Avg('rfx_response_quality'),
            avg_credit_policy=Avg('credit_policy'),
        )

        # Rang du fournisseur
        rank_map = {s['id']: s['rank'] for s in suppliers_with_weighted}
        selected_rank = rank_map.get(selected_supplier.id)

        selected_supplier_data = {
            'id': selected_supplier.id,
            'name': selected_supplier.nom_complet_organisation,
            # Notes globales
            'weighted_rating': float(selected_supplier.get_weighted_rating()),
            'avg_vendor_rating': float(selected_supplier.get_vendor_avg_rating()),
            'avg_buyer_rating': float(selected_supplier.get_buyer_avg_rating()),
            # Compteurs
            'vendor_eval_count': vendor_agg['num'] or 0,
            'buyer_eval_count': buyer_agg['num'] or 0,
            'total_eval_count': (vendor_agg['num'] or 0) + (buyer_agg['num'] or 0),
            # Critères vendor
            'avg_delivery_compliance': (vendor_agg['avg_delivery_compliance'] or 0),
            'avg_delivery_timeline': (vendor_agg['avg_delivery_timeline'] or 0),
            'avg_advising_capability': (vendor_agg['avg_advising_capability'] or 0),
            'avg_after_sales_qos': (vendor_agg['avg_after_sales_qos'] or 0),
            'avg_vendor_relationship': (vendor_agg['avg_vendor_relationship'] or 0),
            # Critères buyer
            'avg_price_flexibility': (buyer_agg['avg_price_flexibility'] or 0),
            'avg_rfx_deadline_compliance': (buyer_agg['avg_rfx_deadline_compliance'] or 0),
            'avg_buyer_advisory_capability': (buyer_agg['avg_advisory_capability'] or 0),
            'avg_relationship_quality': (buyer_agg['avg_relationship_quality'] or 0),
            'avg_rfx_response_quality': (buyer_agg['avg_rfx_response_quality'] or 0),
            'avg_credit_policy': (buyer_agg['avg_credit_policy'] or 0),
            # Rang
            'rank': selected_rank,
        }

        # Yearly breakdown (vendor evaluations)
        yearly_qs = vendor_evals.annotate(year=ExtractYear('date_evaluation')).values('year').annotate(
            avg_delivery_compliance=Avg('delivery_compliance'),
            avg_delivery_timeline=Avg('delivery_timeline'),
            avg_advising_capability=Avg('advising_capability'),
            avg_after_sales_qos=Avg('after_sales_qos'),
            avg_vendor_relationship=Avg('vendor_relationship'),
            avg_final_rating=Avg('vendor_final_rating'),
            num_evaluations=Count('id'),
        ).order_by('year')

        yearly_stats_list = list(yearly_qs)
        yearly_stats_json = yearly_stats_list  # serializable dicts

        # Yearly breakdown (buyer evaluations)
        buyer_yearly_qs = buyer_evals.annotate(year=ExtractYear('date_evaluation')).values('year').annotate(
            avg_price_flexibility=Avg('price_flexibility'),
            avg_rfx_deadline_compliance=Avg('rfx_deadline_compliance'),
            avg_advisory_capability=Avg('advisory_capability'),
            avg_relationship_quality=Avg('relationship_quality'),
            avg_rfx_response_quality=Avg('rfx_response_quality'),
            avg_credit_policy=Avg('credit_policy'),
            avg_final_rating=Avg('buyer_final_rating'),
            num_evaluations=Count('id'),
        ).order_by('year')

        yearly_buyer_stats_list = list(buyer_yearly_qs)
        yearly_buyer_stats_json = yearly_buyer_stats_list

        # Build weighted yearly averages (union of years in vendor and buyer)
        vendor_year_map = {int(y['year']): y for y in yearly_stats_list}
        buyer_year_map = {int(y['year']): y for y in yearly_buyer_stats_list}
        all_years = sorted(set(vendor_year_map.keys()) | set(buyer_year_map.keys()))
        yearly_weighted_json = []
        for y in all_years:
            v_avg = float(vendor_year_map.get(y, {}).get('avg_final_rating') or 0)
            b_avg = float(buyer_year_map.get(y, {}).get('avg_final_rating') or 0)
            # If one side missing for a year, weight the available part accordingly
            weighted = (v_avg * 0.60) + (b_avg * 0.40)
            yearly_weighted_json.append({'year': y, 'weighted_avg': round(weighted, 2), 'vendor_avg': v_avg, 'buyer_avg': b_avg})

        # Per-evaluation time series (ordered by date asc) - vendor evaluations
        evals_ordered = list(vendor_evals.order_by('date_evaluation'))
        running_sum = 0.0
        for idx, e in enumerate(evals_ordered, start=1):
            chart_eval_labels.append(e.date_evaluation.strftime('%d/%m/%Y') if e.date_evaluation else f'#{idx}')
            v = float(e.vendor_final_rating)
            chart_eval_values.append(v)
            running_sum += v
            chart_eval_mavg.append(round(running_sum / idx, 2))

        # Table data (latest first) - vendor evaluations
        evals_for_table = vendor_evals.select_related('evaluator').order_by('-date_evaluation')

    context = {
        'total_suppliers': total_suppliers,
        'suppliers_stats': suppliers_stats,
        'selected_supplier': selected_supplier.nom_complet_organisation if selected_supplier else None,
        'selected_supplier_id': selected_supplier.id if selected_supplier else None,
        'selected_supplier_data': selected_supplier_data,
        'yearly_stats_list': yearly_stats_list,
        'yearly_stats_json': yearly_stats_json,
        'yearly_buyer_stats_list': yearly_buyer_stats_list,
        'yearly_buyer_stats_json': yearly_buyer_stats_json,
        'yearly_weighted_json': yearly_weighted_json,
        'chart_eval_labels': chart_eval_labels,
        'chart_eval_values': chart_eval_values,
        'chart_eval_mavg': chart_eval_mavg,
        'evals_for_table': evals_for_table,
        'top10': top10,
        'bottom10': bottom10,
    }
    return render(request, 'evaluations/ranking_overview.html', context)



@login_required
def export_ranking_top_csv(request):
    """Export Top 10 best suppliers by average rating as CSV"""
    suppliers_qs = Supplier.objects.annotate(
        avg_rating=Avg('evaluations__vendor_final_rating'),
        eval_count=Count('evaluations')
    ).filter(eval_count__gt=0).order_by('-avg_rating', 'nom_complet_organisation')[:10]

    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="ranking_top10.csv"'
    import csv
    w = csv.writer(resp)
    w.writerow(['Rank', 'Supplier', 'Average', '#Evaluations'])
    for idx, s in enumerate(suppliers_qs, start=1):
        w.writerow([idx, s.nom_complet_organisation, f"{(s.avg_rating or 0):.2f}", s.eval_count])
    return resp


@login_required
def export_ranking_xlsx(request):
    """Export an XLSX workbook containing Top10, Bottom10, overview and optional selected supplier details."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        return HttpResponse("openpyxl is required for XLSX export. Please install it: pip install openpyxl", status=500)

    # Prepare data
    suppliers_qs = Supplier.objects.annotate(
        avg_rating=Avg('evaluations__vendor_final_rating'),
        eval_count=Count('evaluations')
    ).filter(eval_count__gt=0)

    top10 = list(suppliers_qs.order_by('-avg_rating', 'nom_complet_organisation').values('nom_complet_organisation', 'avg_rating', 'eval_count')[:10])
    bottom10 = list(suppliers_qs.order_by('avg_rating', 'nom_complet_organisation').values('nom_complet_organisation', 'avg_rating', 'eval_count')[:10])

    supplier_id = request.GET.get('supplier')
    include_yearly = request.GET.get('yearly') in ['1', 'true', 'on', 'True']
    include_chart = request.GET.get('chart') in ['1', 'true', 'on', 'True']
    supplier = None
    supplier_evals = []
    if supplier_id:
        supplier = get_object_or_404(Supplier, pk=supplier_id)
        supplier_evals = list(
            SupplierEvaluation.objects.filter(supplier=supplier)
            .select_related('evaluator')
            .order_by('date_evaluation')
            .values('date_evaluation','vendor_final_rating','delivery_compliance','delivery_timeline','advising_capability','after_sales_qos','vendor_relationship','evaluator__email','comments')
        )

    # Build workbook
    wb = openpyxl.Workbook()
    ws_over = wb.active
    ws_over.title = 'Overview'
    ws_top = wb.create_sheet(title='Top10')
    ws_bottom = wb.create_sheet(title='Bottom10')
    ws_supplier = wb.create_sheet(title='Supplier')
    ws_yearly = wb.create_sheet(title='Yearly') if include_yearly else None

    header_fill = PatternFill(start_color='FFDDE9FF', end_color='FFDDE9FF', fill_type='solid')  # light blue
    bold = Font(bold=True)
    thin = Side(border_style='thin', color='FFCCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')

    # Overview
    ws_over.append(['Metric','Value'])
    ws_over['A1'].font = bold; ws_over['B1'].font = bold
    ws_over['A1'].fill = header_fill; ws_over['B1'].fill = header_fill
    total_suppliers = suppliers_qs.count()
    overall_avg = suppliers_qs.aggregate(a=Avg('avg_rating'))['a'] or 0
    ws_over.append(['Suppliers evaluated', total_suppliers])
    ws_over.append(['Global average (avg of supplier avgs)', round(overall_avg,2)])
    for row in ws_over.iter_rows(min_row=1, max_row=ws_over.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = border

    # Top10
    ws_top.append(['Rank','Supplier','Average','Evaluations'])
    for cell in ws_top[1]:
        cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border
    for idx, s in enumerate(top10, start=1):
        ws_top.append([idx, s['nom_complet_organisation'], float(s['avg_rating'] or 0), int(s['eval_count'] or 0)])
    for row in ws_top.iter_rows(min_row=2, max_row=ws_top.max_row, min_col=1, max_col=4):
        for cell in row: cell.border = border

    # Bottom10
    ws_bottom.append(['Rank','Supplier','Average','Evaluations'])
    for cell in ws_bottom[1]:
        cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border
    for idx, s in enumerate(bottom10, start=1):
        ws_bottom.append([idx, s['nom_complet_organisation'], float(s['avg_rating'] or 0), int(s['eval_count'] or 0)])
    for row in ws_bottom.iter_rows(min_row=2, max_row=ws_bottom.max_row, min_col=1, max_col=4):
        for cell in row: cell.border = border

    # Supplier details
    ws_supplier.append(['Supplier'])
    ws_supplier.append([supplier.nom_complet_organisation if supplier else '—'])
    ws_supplier.append([])
    ws_supplier.append(['Date','Final','Delivery','Timeline','Advising','After Sales','Relationship','Evaluator','Comments'])
    for cell in ws_supplier[4]:
        cell.font = bold; cell.fill = header_fill; cell.border = border
    for e in supplier_evals:
        ws_supplier.append([
            e['date_evaluation'].strftime('%Y-%m-%d') if e['date_evaluation'] else '',
            float(e['vendor_final_rating'] or 0),
            e['delivery_compliance'],
            e['delivery_timeline'],
            e['advising_capability'],
            e['after_sales_qos'],
            e['vendor_relationship'],
            e.get('evaluator__email') or '',
            (e.get('comments') or '').replace('\n',' ').strip(),
        ])
    for row in ws_supplier.iter_rows(min_row=4, max_row=ws_supplier.max_row, min_col=1, max_col=9):
        for cell in row: cell.border = border

    # Yearly sheet for selected supplier
    if include_yearly and supplier:
        from django.db.models.functions import ExtractYear
        yearly_qs = SupplierEvaluation.objects.filter(supplier=supplier).annotate(year=ExtractYear('date_evaluation')).values('year').annotate(
            avg_delivery_compliance=Avg('delivery_compliance'),
            avg_delivery_timeline=Avg('delivery_timeline'),
            avg_advising_capability=Avg('advising_capability'),
            avg_after_sales_qos=Avg('after_sales_qos'),
            avg_vendor_relationship=Avg('vendor_relationship'),
            avg_final_rating=Avg('vendor_final_rating'),
            num_evaluations=Count('id'),
        ).order_by('year')
        ws_yearly.append(['Year','Avg Final','Delivery','Timeline','Advising','After Sales','Relationship','#Evals'])
        for cell in ws_yearly[1]:
            cell.font = bold; cell.fill = header_fill; cell.alignment = center; cell.border = border
        for y in yearly_qs:
            ws_yearly.append([
                y['year'],
                float(y['avg_final_rating'] or 0),
                float(y['avg_delivery_compliance'] or 0),
                float(y['avg_delivery_timeline'] or 0),
                float(y['avg_advising_capability'] or 0),
                float(y['avg_after_sales_qos'] or 0),
                float(y['avg_vendor_relationship'] or 0),
                int(y['num_evaluations'] or 0),
            ])
        for row in ws_yearly.iter_rows(min_row=2, max_row=ws_yearly.max_row, min_col=1, max_col=8):
            for cell in row: cell.border = border

        if include_chart and ws_yearly.max_row >= 2:
            # Create a line chart for Avg Final by Year
            chart = LineChart()
            chart.title = "Average Final Rating by Year"
            data = Reference(ws_yearly, min_col=2, min_row=1, max_col=2, max_row=ws_yearly.max_row)
            cats = Reference(ws_yearly, min_col=1, min_row=2, max_row=ws_yearly.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 10
            ws_yearly.add_chart(chart, "J2")

    # Autosize columns
    for ws in [ws_over, ws_top, ws_bottom, ws_supplier]:
        for col in ws.columns:
            maxlen = 10
            for cell in col:
                try:
                    maxlen = max(maxlen, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(maxlen + 2, 60)

    # Response
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="supplier_ranking.xlsx"'
    wb.save(resp)
    return resp


@login_required
def export_ranking_bottom_csv(request):
    """Export Top 10 worst suppliers by average rating as CSV"""
    suppliers_qs = Supplier.objects.annotate(
        avg_rating=Avg('evaluations__vendor_final_rating'),
        eval_count=Count('evaluations')
    ).filter(eval_count__gt=0).order_by('avg_rating', 'nom_complet_organisation')[:10]

    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="ranking_bottom10.csv"'
    import csv
    w = csv.writer(resp)
    w.writerow(['Rank', 'Supplier', 'Average', '#Evaluations'])
    for idx, s in enumerate(suppliers_qs, start=1):
        w.writerow([idx, s.nom_complet_organisation, f"{(s.avg_rating or 0):.2f}", s.eval_count])
    return resp


@login_required
def export_supplier_ranking_csv(request):
    """Export selected supplier evaluations detail as CSV"""
    supplier_id = request.GET.get('supplier')
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    evals = SupplierEvaluation.objects.filter(supplier=supplier).select_related('evaluator').order_by('date_evaluation')

    resp = HttpResponse(content_type='text/csv')
    safe_name = supplier.nom_complet_organisation.replace(' ', '_')
    resp['Content-Disposition'] = f'attachment; filename="supplier_ranking_{safe_name}.csv"'
    import csv
    w = csv.writer(resp)
    w.writerow(['Date', 'Final Rating', 'Delivery Compliance', 'Timeline', 'Advising', 'After Sales', 'Relationship', 'Evaluator', 'Comments'])
    for e in evals:
        w.writerow([
            e.date_evaluation.strftime('%Y-%m-%d') if e.date_evaluation else '',
            f"{e.vendor_final_rating}",
            e.delivery_compliance,
            e.delivery_timeline,
            e.advising_capability,
            e.after_sales_qos,
            e.vendor_relationship,
            getattr(e.evaluator, 'email', '') or '',
            (e.comments or '').replace('\n', ' ').strip(),
        ])
    return resp


# ============================================
# VUES POUR ÉVALUATION ACHETEUR
# ============================================

@login_required
def buyer_evaluation_list(request):
    """Liste des évaluations acheteur"""
    evaluations = BuyerEvaluation.objects.select_related('supplier', 'evaluator').all()
    
    # Filtrage par fournisseur
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        evaluations = evaluations.filter(supplier_id=supplier_id)
    
    # Recherche
    search = request.GET.get('search')
    if search:
        evaluations = evaluations.filter(
            Q(supplier__nom_complet_organisation__icontains=search) |
            Q(evaluator__email__icontains=search) |
            Q(comments__icontains=search)
        )
    
    # Statistiques
    stats = {
        'total': evaluations.count(),
        'avg_rating': evaluations.aggregate(Avg('buyer_final_rating'))['buyer_final_rating__avg'] or 0,
    }
    
    # Liste des fournisseurs pour le filtre
    suppliers = Supplier.objects.filter(actif=True).order_by('nom_complet_organisation')
    
    context = {
        'evaluations': evaluations,
        'suppliers': suppliers,
        'stats': stats,
    }
    return render(request, 'evaluations/buyer_evaluation_list.html', context)


@login_required
def buyer_evaluation_detail(request, pk):
    """Détail d'une évaluation acheteur"""
    evaluation = get_object_or_404(BuyerEvaluation, pk=pk)
    
    # Récupérer toutes les évaluations acheteur du même fournisseur
    supplier_evaluations = BuyerEvaluation.objects.filter(
        supplier=evaluation.supplier
    ).exclude(pk=pk).order_by('-date_evaluation')[:5]
    
    # Préparer les descriptions des critères pour le template
    descriptions = {
        'price_flexibility': evaluation.get_criteria_description('price_flexibility', evaluation.price_flexibility),
        'rfx_deadline_compliance': evaluation.get_criteria_description('rfx_deadline_compliance', evaluation.rfx_deadline_compliance),
        'advisory_capability': evaluation.get_criteria_description('advisory_capability', evaluation.advisory_capability),
        'relationship_quality': evaluation.get_criteria_description('relationship_quality', evaluation.relationship_quality),
        'rfx_response_quality': evaluation.get_criteria_description('rfx_response_quality', evaluation.rfx_response_quality),
        'credit_policy': evaluation.get_criteria_description('credit_policy', evaluation.credit_policy),
    }
    
    context = {
        'evaluation': evaluation,
        'supplier_evaluations': supplier_evaluations,
        'descriptions': descriptions,
    }
    return render(request, 'evaluations/buyer_evaluation_detail.html', context)


@login_required
def buyer_evaluation_create(request):
    """Créer une évaluation acheteur"""
    if request.method == 'POST':
        form = BuyerEvaluationForm(request.POST)
        if form.is_valid():
            evaluation = form.save(commit=False)
            evaluation.evaluator = request.user
            evaluation.save()
            messages.success(request, f'Évaluation acheteur du fournisseur "{evaluation.supplier.nom_complet_organisation}" créée avec succès !')
            return redirect('evaluations:buyer_detail', pk=evaluation.pk)
    else:
        # Pré-remplir le fournisseur si passé en paramètre
        supplier_id = request.GET.get('supplier')
        initial = {}
        if supplier_id:
            initial['supplier'] = supplier_id
        form = BuyerEvaluationForm(initial=initial)
    
    context = {'form': form}
    return render(request, 'evaluations/buyer_evaluation_form.html', context)


@login_required
def buyer_evaluation_edit(request, pk):
    """Modifier une évaluation acheteur"""
    evaluation = get_object_or_404(BuyerEvaluation, pk=pk)
    
    if request.method == 'POST':
        form = BuyerEvaluationForm(request.POST, instance=evaluation)
        if form.is_valid():
            evaluation = form.save()
            messages.success(request, f'Évaluation acheteur modifiée avec succès !')
            return redirect('evaluations:buyer_detail', pk=evaluation.pk)
    else:
        form = BuyerEvaluationForm(instance=evaluation)
    
    context = {'form': form, 'evaluation': evaluation}
    return render(request, 'evaluations/buyer_evaluation_form.html', context)


@login_required
def buyer_evaluation_delete(request, pk):
    """Supprimer une évaluation acheteur"""
    evaluation = get_object_or_404(BuyerEvaluation, pk=pk)
    
    if request.method == 'POST':
        supplier_name = evaluation.supplier.nom_complet_organisation
        evaluation.delete()
        messages.success(request, f'Évaluation acheteur du fournisseur "{supplier_name}" supprimée avec succès.')
        return redirect('evaluations:buyer_list')
    
    # Si GET, rediriger vers la liste
    return redirect('evaluations:buyer_list')


@login_required
def supplier_buyer_evaluations(request, supplier_id):
    """Liste des évaluations acheteur d'un fournisseur spécifique"""
    supplier = get_object_or_404(Supplier, pk=supplier_id)
    evaluations = BuyerEvaluation.objects.filter(supplier=supplier).order_by('-date_evaluation')
    
    # Statistiques du fournisseur
    stats = {
        'total': evaluations.count(),
        'avg_rating': evaluations.aggregate(Avg('buyer_final_rating'))['buyer_final_rating__avg'] or 0,
        'avg_price_flexibility': evaluations.aggregate(Avg('price_flexibility'))['price_flexibility__avg'] or 0,
        'avg_rfx_deadline_compliance': evaluations.aggregate(Avg('rfx_deadline_compliance'))['rfx_deadline_compliance__avg'] or 0,
        'avg_advisory_capability': evaluations.aggregate(Avg('advisory_capability'))['advisory_capability__avg'] or 0,
        'avg_relationship_quality': evaluations.aggregate(Avg('relationship_quality'))['relationship_quality__avg'] or 0,
        'avg_rfx_response_quality': evaluations.aggregate(Avg('rfx_response_quality'))['rfx_response_quality__avg'] or 0,
        'avg_credit_policy': evaluations.aggregate(Avg('credit_policy'))['credit_policy__avg'] or 0,
    }
    
    context = {
        'supplier': supplier,
        'evaluations': evaluations,
        'stats': stats,
    }
    return render(request, 'evaluations/supplier_buyer_evaluations.html', context)
